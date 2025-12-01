#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspect XLSX columns to find where the actual poetry is
"""
import openpyxl

wb = openpyxl.load_workbook('GLC_Anemoskala_select_text.xlsx')
ws = wb.active

print(f"Total columns: {ws.max_column}")
print(f"Total rows: {ws.max_row}")

# Check first row to see headers
print("\n=== First Row (Headers?) ===")
first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
for idx, cell in enumerate(first_row, start=1):
    print(f"Column {idx}: {cell}")

# Sample a few rows to see content
print("\n=== Sample Rows 2-10 ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=10, values_only=True), start=2):
    print(f"\nRow {row_idx}:")
    for col_idx, cell in enumerate(row, start=1):
        if cell:
            cell_str = str(cell)[:100]
            print(f"  Col {col_idx}: {cell_str}")
